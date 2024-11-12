##extraction_script_ver2
##크롤링 게시글 개수에 대해 사용자에게 입력받는 코드 추가

import sys
import time
import traceback
import logging
from typing import Optional, List, Dict

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# 크롤링할 최대 게시글 수 설정
MAX_POSTS = 20
EXCEL_FILE = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'  # 엑셀 파일 경로
WORKSHEET_NAME = '개인정보 추출 및 이용 관리'

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def initialize_webdriver() -> webdriver.Chrome:
    """
    웹드라이버 초기화
    """
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    return driver


def login(driver: webdriver.Chrome, username: str, password: str) -> bool:
    """
    로그인 처리
    """
    try:
        # 로그인 페이지로 이동
        driver.get('https://gw.com2us.com/')
        
        # 로그인 폼 요소 찾기
        username_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'Username'))
        )
        password_input = driver.find_element(By.ID, 'Password')

        # 로그인 정보 입력 및 제출
        username_input.send_keys(username)
        password_input.send_keys(password)
        driver.find_element(By.CLASS_NAME, 'btnLogin').click()

        # 로그인 성공 여부 확인
        WebDriverWait(driver, 30).until(EC.url_changes('https://gw.com2us.com/'))
        if 'login' in driver.current_url.lower():
            logging.error("로그인에 실패하였습니다.")
            return False
        return True
    except Exception as e:
        logging.error(f"로그인 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def navigate_to_search_page(driver: webdriver.Chrome) -> bool:
    """
    검색 페이지로 이동하는 함수
    """
    try:
        # 검색 페이지 URL로 이동
        driver.get('https://gw.com2us.com/emate_appro/appro_complete_2024_link.nsf/wfmViaView?readform&viewname=view055&vctype=a')
        # 검색창이 로드될 때까지 대기
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'searchtext')))
        return True
    except Exception as e:
        logging.error(f"검색 페이지 이동 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def search_documents(driver: webdriver.Chrome) -> bool:
    """
    '개인정보 추출 신청서' 검색을 수행하는 함수
    """
    try:
        # 검색어 입력
        search_input = driver.find_element(By.ID, 'searchtext')
        search_input.clear()
        search_input.send_keys('개인정보 추출 신청서')
        
        # 검색 버튼 클릭
        search_button = driver.find_element(By.XPATH, '//img[@class="inbtn" and contains(@src, "btn_search_board.gif")]')
        search_button.click()
        time.sleep(5)
        return True
    except Exception as e:
        logging.error(f"문서 검색 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def extract_post_data(driver: webdriver.Chrome, post: webdriver.remote.webelement.WebElement, index: int) -> Optional[Dict]:
    """
    개별 게시글의 데이터를 추출하는 함수
    """
    try:
        # 기본 정보 추출
        tds = post.find_elements(By.TAG_NAME, 'td')
        
        # 결재일 추출
        결재일_text = tds[5].text.strip()
        년, 월, 일 = 결재일_text.split('-')
        월 = str(int(월))
        일 = str(int(일))
        
        # 신청자 추출
        신청자 = tds[4].find_element(By.TAG_NAME, 'span').text.strip()

        # 게시글 상세 페이지 열기
        driver.execute_script("arguments[0].scrollIntoView();", post)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable(post))
        post.click()

        # 새 창으로 전환
        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
        driver.switch_to.window(driver.window_handles[-1])

        # 상세 페이지 데이터 추출
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'AppLineArea')))
        
        # 문서 종류 확인
        h2_element = driver.find_element(By.CSS_SELECTOR, '#AppLineArea h2')
        if h2_element.text.strip() != '개인정보 추출 신청서':
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            return None

        # 상세 정보 추출
        법인명_elements = driver.find_elements(By.ID, 'titleLabel')
        법인명 = 법인명_elements[0].text.strip() if 법인명_elements else ''

        문서번호_elements = driver.find_elements(By.XPATH, '//th[contains(text(),"문서번호")]/following-sibling::td[1]')
        문서번호 = 문서번호_elements[0].text.strip() if 문서번호_elements else ''

        # 제목_elements = driver.find_elements(By.CSS_SELECTOR, 'td.approval_text')
        # 제목 = 제목_elements[0].text.strip() if 제목_elements else ''
        제목_elements = driver.find_elements(By.CSS_SELECTOR, 'td.approval_text')
        if 제목_elements:
            제목_text = 제목_elements[0].text.strip()
            # 제목이 법인명 + 공백으로 시작하면, 그 부분을 제거
            if 제목_text.startswith(법인명 + ' '):
                제목 = 제목_text[len(법인명) + 1:].strip()
            else:
                제목 = 제목_text
        else:
            제목 = ''

        합의담당자_elements = driver.find_elements(By.XPATH, '//th[text()="합의선"]/following::tr[@class="name"][1]/td[@class="td_point"]')
        합의담당자 = 합의담당자_elements[0].text.strip() if 합의담당자_elements else ''

        # 추출 데이터 구성
        data = {
            '결재일': 결재일_text,
            '년': 년,
            '월': 월,
            '일': 일,
            '주차': '',
            '법인명': 법인명,
            '문서번호': 문서번호,
            '제목': 제목,
            '업무 유형': '',
            '추출 위치': '',
            '담당 부서': '',
            '신청자': 신청자,
            '합의 담당자': 합의담당자,
            '링크': driver.current_url,
            '진행 구분': ''
        }

        logging.info(f"게시글 {index}: 데이터 추출 완료")
        return data

    except Exception as e:
        logging.error(f"게시글 {index}: 데이터 추출 중 오류 발생: {e}")
        traceback.print_exc()
        return None
    finally:
        try:
            window_handles = driver.window_handles
            if len(window_handles) > 1:
                # 현재 창(팝업 창)을 닫고, 메인 창으로 전환
                driver.close()
                driver.switch_to.window(window_handles[0])
            else:
                # 창이 하나밖에 없을 경우, 전환만 수행
                driver.switch_to.window(window_handles[0])
            time.sleep(2)
        except Exception as e:
            logging.error(f"창 전환 중 오류 발생: {e}")
            traceback.print_exc()


def save_to_excel(data_list: List[Dict]) -> None:
    """
    추출된 데이터를 엑셀 파일에 저장하는 함수
    """
    if not data_list:
        logging.info("추출된 데이터가 없습니다.")
        return

    try:
        # 엑셀 파일 열기
        wb = load_workbook(EXCEL_FILE)
        if WORKSHEET_NAME not in wb.sheetnames:
            logging.error(f"워크시트 '{WORKSHEET_NAME}'이(가) 존재하지 않습니다.")
            return
        ws = wb[WORKSHEET_NAME]

        # 마지막 행 찾기
        last_row = ws.max_row
        while last_row >= 6:
            if ws.cell(row=last_row, column=2).value is not None:
                break
            last_row -= 1

        # 시작 행과 번호 설정
        start_row = max(6, last_row + 1)
        if last_row >= 6:
            last_no = ws.cell(row=last_row, column=2).value
            next_no = last_no + 1 if isinstance(last_no, int) else 1
        else:
            next_no = 1

        # 엑셀 열 매핑 정의
        column_mapping = {
            '결재일': 3, '년': 4, '월': 5, '일': 6, '주차': 7,
            '법인명': 8, '문서번호': 9, '제목': 10, '업무 유형': 11,
            '추출 위치': 12, '담당 부서': 13, '신청자': 14,
            '합의 담당자': 15, '링크': 16, '진행 구분': 17
        }

        # 데이터 쓰기
        for data in data_list:
            ws.cell(row=start_row, column=2, value=next_no)
            next_no += 1
            for col_name, col_idx in column_mapping.items():
                ws.cell(row=start_row, column=col_idx, value=data[col_name])
            start_row += 1

        # 파일 저장
        wb.save(EXCEL_FILE)
        logging.info(f"데이터가 성공적으로 '{EXCEL_FILE}' 파일에 저장되었습니다.")

    except Exception as e:
        logging.error(f"엑셀 저장 중 오류 발생: {e}")
        traceback.print_exc()

def main(username: str, password: str, max_posts: Optional[int] = None) -> Optional[str]:
    driver = initialize_webdriver()

    try:
        # 로그인 및 페이지 이동
        if not login(driver, username, password):
            driver.quit()
            return None

        if not navigate_to_search_page(driver):
            driver.quit()
            return None

        if not search_documents(driver):
            driver.quit()
            return None

        # 게시글 데이터 추출
        data_list = []
        posts = driver.find_elements(By.XPATH, '//tr[contains(@class, "dhx_skyblue")]')

        # 최대 게시글 수 설정
        if max_posts is not None:
            num_posts_to_crawl = min(len(posts), max_posts)
        else:
            num_posts_to_crawl = len(posts)

        for i in range(num_posts_to_crawl):
            posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
            if i >= len(posts):
                logging.warning(f"게시글 수가 예상보다 적습니다. 현재 인덱스: {i}, 게시글 수: {len(posts)}")
                break

            data = extract_post_data(driver, posts[i], i + 1)
            if data:
                data_list.append(data)

        # 추출된 데이터 저장
        save_to_excel(data_list)

        # 완료 메시지와 파일 경로 반환
        logging.info("개인정보 신청 이력 저장이 완료되었습니다.")
        return EXCEL_FILE

    except Exception as e:
        logging.error(f"메인 프로세스 중 오류 발생: {e}")
        traceback.print_exc()
        return None
    finally:
        driver.quit()
        logging.info("브라우저가 종료되었습니다.")
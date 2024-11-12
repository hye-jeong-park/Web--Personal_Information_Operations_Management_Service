## 저장할 파일 선택하고(구현 완료), 링크 통해서 파일 다운로드하는 방법(추가 구현 필요) ##
import re
import sys
import time
import traceback
import logging
from typing import Tuple, Optional, List, Dict

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def initialize_webdriver() -> webdriver.Chrome:
    """
    웹드라이버 초기화
    """
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    return driver


def login(driver: webdriver.Chrome, username: str, password: str) -> bool:
    """
    로그인 처리
    """
    try:
        driver.get('https://gw.com2us.com/')
        username_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'Username'))
        )
        password_input = driver.find_element(By.ID, 'Password')

        username_input.send_keys(username)
        password_input.send_keys(password)
        driver.find_element(By.CLASS_NAME, 'btnLogin').click()

        WebDriverWait(driver, 20).until(EC.url_changes('https://gw.com2us.com/'))
        if 'login' in driver.current_url.lower():
            logging.error("로그인에 실패하였습니다.")
            return False
        return True
    except Exception as e:
        logging.error(f"로그인 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def navigate_to_target_page(driver: webdriver.Chrome) -> bool:
    """
    개인정보 파일 전송 페이지로 이동
    """
    try:
        driver.get('https://gw.com2us.com/emate_app/00001/bbs/b2307140306.nsf/view?readform&viewname=view01')
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]'))
        )
        logging.info(f"페이지 이동 후 현재 URL: {driver.current_url}")
        return True
    except Exception as e:
        logging.error(f"타겟 페이지로 이동 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def fetch_posts(driver: webdriver.Chrome) -> List[webdriver.remote.webelement.WebElement]:
    """
    현재 페이지의 게시글 목록을 가져옵니다.
    """
    try:
        posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
        total_posts = len(posts)
        logging.info(f"총 게시글 수: {total_posts}")
        return posts
    except Exception as e:
        logging.error("게시글 목록을 가져오는 중 오류 발생.")
        logging.error(e)
        traceback.print_exc()
        return []


def go_to_page(driver: webdriver.Chrome, page_number: int) -> bool:
    """
    주어진 페이지 번호로 이동합니다.
    """
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'pagingNav'))
        )

        current_page_element = driver.find_element(By.CSS_SELECTOR, 'div#pagingNav strong.cur_num')
        current_page = int(current_page_element.text.strip())
        if current_page == page_number:
            return True

        page_links = driver.find_elements(By.XPATH, f'//div[@id="pagingNav"]//a[@class="num_box"]')
        page_link = None
        for link in page_links:
            if link.text.strip() == str(page_number):
                page_link = link
                break

        if page_link:
            page_link.click()
        else:
            logging.info(f"페이지 번호 {page_number}를 찾을 수 없습니다.")
            return False

        time.sleep(2)
        WebDriverWait(driver, 10).until(
            EC.text_to_be_present_in_element((By.CSS_SELECTOR, 'div#pagingNav strong.cur_num'), str(page_number))
        )
        logging.info(f"{page_number} 페이지로 이동 완료")
        return True
    except Exception as e:
        logging.error(f"{page_number} 페이지로 이동 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def extract_corporate_name(full_text: str) -> str:
    """
    법인명 추출
    """
    if '/' in full_text:
        return full_text.split('/')[0].strip().split()[0]
    return full_text.strip().split()[0]


def extract_post_data(driver: webdriver.Chrome, post: webdriver.remote.webelement.WebElement, index: int) -> Optional[Dict]:
    """
    개별 게시글의 데이터를 추출하는 함수
    """
    try:
        tds = post.find_elements(By.TAG_NAME, 'td')
        등록일_text = tds[4].get_attribute('title').strip() if tds[4].get_attribute('title') else tds[4].text.strip()
        작성자_td = tds[2]
        작성자 = 작성자_td.find_element(By.TAG_NAME, 'span').text.strip() if 작성자_td.find_elements(By.TAG_NAME, 'span') else 작성자_td.text.strip()

        driver.execute_script("arguments[0].scrollIntoView();", post)
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable(post))
        post.click()

        WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(2))
        driver.switch_to.window(driver.window_handles[-1])
        logging.info(f"게시글 {index}: 새 창으로 전환")

        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'HeaderTable')))
        제목 = driver.find_element(By.ID, 'DisSubject').text.strip() if driver.find_elements(By.ID, 'DisSubject') else ''
        작성자_full = driver.find_element(By.ID, 'DismyName').text.strip() if driver.find_elements(By.ID, 'DismyName') else ''
        등록일_text_detail = driver.find_element(By.ID, 'DiscDate').text.strip() if driver.find_elements(By.ID, 'DiscDate') else ''

        data = {
            '등록일': 등록일_text or 등록일_text_detail,
            '법인명': extract_corporate_name(작성자_full),
            '제목': 제목,
            '작성자': 작성자_full,
            '링크': driver.current_url
        }

        logging.info(f"게시글 {index}: 데이터 추출 완료")
        return data

    except Exception as e:
        logging.error(f"게시글 {index}: 데이터 추출 중 오류 발생: {e}")
        traceback.print_exc()
        return None
    finally:
        try:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(2)
        except Exception as e:
            logging.error(f"창 전환 중 오류 발생: {e}")
            traceback.print_exc()


def save_to_excel(data_list: List[Dict], excel_path: str) -> None:
    """
    추출된 데이터를 엑셀 파일에 저장하는 함수
    """
    if not data_list:
        logging.info("추출된 데이터가 없습니다.")
        return

    try:
        wb = load_workbook(excel_path)
        ws = wb['개인정보 추출 및 이용 관리']
        max_row = ws.max_row

        column_mapping = {
            '등록일': 19,           # S
            '법인명': 20,           # T
            '제목': 21,             # U
            '작성자': 22,           # V
            '링크': 23              # W
        }

        for data in data_list:
            found_row = None
            for row_idx in range(6, max_row + 1):
                if ws.cell(row=row_idx, column=16).value == data['링크']:
                    found_row = row_idx
                    break

            if found_row:
                for col_name, col_idx in column_mapping.items():
                    ws.cell(row=found_row, column=col_idx, value=data[col_name])
                logging.info(f"데이터가 엑셀의 행 {found_row}에 저장되었습니다.")
            else:
                logging.warning(f"링크 '{data['링크']}'를 가진 행을 찾을 수 없습니다.")

        wb.save(excel_path)
        logging.info(f"데이터가 성공적으로 '{excel_path}' 파일에 저장되었습니다.")

    except Exception as e:
        logging.error(f"엑셀 파일 처리 중 오류 발생: {e}")
        traceback.print_exc()


def main(username: str, password: str, max_posts: Optional[int], excel_path: str) -> Optional[str]:
    driver = initialize_webdriver()
    try:
        if not login(driver, username, password):
            return None
        if not navigate_to_target_page(driver):
            return None

        data_list = []
        total_crawled = 0
        page_number = 1

        while True:
            logging.info(f"{page_number} 페이지 크롤링 시작")
            posts = fetch_posts(driver)
            if not posts:
                break

            for i in range(1 if page_number == 1 else 0, len(posts)):
                if max_posts and total_crawled >= max_posts:
                    break
                data = extract_post_data(driver, posts[i], total_crawled + 1)
                if data:
                    data_list.append(data)
                    total_crawled += 1

            if max_posts and total_crawled >= max_posts:
                break

            page_number += 1
            if not go_to_page(driver, page_number):
                break

        save_to_excel(data_list, excel_path)
        logging.info("개인정보 추출 및 전달이 완료되었습니다.")
        return excel_path

    except Exception as e:
        logging.error("스크립트 실행 중 오류 발생: {e}")
        return None
    finally:
        driver.quit()
        logging.info("브라우저가 종료되었습니다.")
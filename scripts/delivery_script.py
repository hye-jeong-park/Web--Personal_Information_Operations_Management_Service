import re
import sys
import time
import traceback
import logging
from typing import Tuple, Optional, List, Dict
import getpass

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

from .extraction_in_progress_script import run_extraction_in_progress

MAX_POSTS = 50
EXCEL_FILE = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
WORKSHEET_NAME = '개인정보 추출 및 이용 관리'

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def initialize_webdriver() -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    return driver


def login(driver: webdriver.Chrome, username: str, password: str) -> bool:
    try:
        driver.get('https://gw.com2us.com/')
        username_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'Username'))
        )
        password_input = driver.find_element(By.ID, 'Password')
        username_input.send_keys(username)
        password_input.send_keys(password)
        driver.find_element(By.CLASS_NAME, 'btnLogin').click()

        WebDriverWait(driver, 20).until(EC.url_changes('https://gw.com2us.com/'))
        if 'login' in driver.current_url.lower():
            logging.error("로그인에 실패하였습니다.")
            return False
        return True
    except Exception as e:
        logging.error(f"로그인 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def navigate_to_target_page(driver: webdriver.Chrome) -> bool:
    try:
        driver.get('https://gw.com2us.com/emate_app/00001/bbs/b2307140306.nsf/view?readform&viewname=view01')
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]'))
        )
        logging.info(f"페이지 이동 후 현재 URL: {driver.current_url}")
        return True
    except Exception as e:
        logging.error(f"타겟 페이지로 이동 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def fetch_posts(driver: webdriver.Chrome) -> List[webdriver.remote.webelement.WebElement]:
    try:
        posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
        total_posts = len(posts)
        logging.info(f"총 게시글 수: {total_posts}")
        return posts
    except Exception as e:
        logging.error("게시글 목록을 가져오는 중 오류 발생.")
        logging.error(e)
        traceback.print_exc()
        return []


def go_to_page(driver: webdriver.Chrome, page_number: int) -> bool:
    try:
        if not driver.session_id:
            logging.error("세션이 만료되었습니다.")
            return False

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'pagingNav'))
        )

        current_page_element = driver.find_element(By.CSS_SELECTOR, 'div#pagingNav strong.cur_num')
        current_page = int(current_page_element.text.strip())

        if current_page == page_number:
            return True

        page_links = driver.find_elements(By.XPATH, f'//div[@id="pagingNav"]//a[@class="num_box"]')
        page_link = None
        for link in page_links:
            if link.text.strip() == str(page_number):
                page_link = link
                break

        if page_link:
            page_link.click()
            time.sleep(2)
            WebDriverWait(driver, 10).until(
                EC.text_to_be_present_in_element((By.CSS_SELECTOR, 'div#pagingNav strong.cur_num'), str(page_number))
            )
            logging.info(f"{page_number} 페이지로 이동 완료")
            return True
        else:
            logging.info(f"페이지 번호 {page_number}를 찾을 수 없습니다.")
            return False

    except Exception as e:
        logging.error(f"{page_number} 페이지로 이동 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def extract_corporate_name(full_text: str) -> str:
    if '/' in full_text:
        return full_text.split('/')[0].strip().split()[0]
    return full_text.strip().split()[0]


def extract_department_name(full_text: str) -> str:
    try:
        if '/' in full_text:
            parts = full_text.split('/')
            if len(parts) >= 2:
                department_full = parts[1].strip()
                department_parts = department_full.split()
                if len(department_parts) >= 1:
                    return department_parts[-1]
        elif ',' in full_text:
            parts = full_text.split(',')
            if len(parts) >= 1:
                department_full = parts[0].strip()
                department_parts = department_full.split()
                if len(department_parts) >= 2:
                    return department_parts[1]
        return ''
    except Exception as e:
        logging.error(f"부서명 추출 중 오류 발생: {e}")
        return ''


def extract_file_info(file_info: str) -> Tuple[str, str]:
    file_match = re.match(r'(.+?)\s*(?:&|[(])\s*([\d,\.]+\s*[KMGT]?B)', file_info, re.IGNORECASE)
    if file_match:
        filename_part = file_match.group(1).strip()
        size_part = file_match.group(2).strip()
    else:
        filename_part = file_info.strip()
        size_match = re.search(r'([\d,\.]+\s*[KMGT]?B)', filename_part, re.IGNORECASE)
        if size_match:
            size_part = size_match.group(1).strip()
            filename_part = filename_part.replace(size_part, '').strip()
        else:
            size_part = ''

    file_type = ''
    if '.zip' in filename_part.lower():
        file_type = 'Zip'
    elif '.xlsx' in filename_part.lower():
        file_type = 'Excel'

    size_match = re.match(r'([\d,\.]+)\s*([KMGT]?B)', size_part, re.IGNORECASE)
    if size_match:
        size_numeric = size_match.group(1).replace(',', '')
        size_unit = size_match.group(2).upper()
        file_size = f"{size_numeric} {size_unit}"
    else:
        file_size = size_part

    return file_type, file_size


def find_section_text(driver: webdriver.Chrome, section_titles: List[str]) -> Optional[str]:
    try:
        tr_elements = driver.find_elements(By.XPATH, '//table//tr')
        for tr in tr_elements:
            try:
                td_elements = tr.find_elements(By.TAG_NAME, 'td')
                if len(td_elements) >= 2:
                    header_text = ''.join([span.text.strip() for span in td_elements[0].find_elements(By.TAG_NAME, 'span')])
                    for section_title in section_titles:
                        if section_title in header_text:
                            return td_elements[1].text.strip()
            except Exception:
                continue
        return None
    except Exception as e:
        logging.error(f"find_section_text 오류: {e}")
        return None


def extract_attachment_info(driver: webdriver.Chrome) -> Tuple[str, str]:
    파일형식, 파일용량 = '', ''

    try:
        attm_read_div = driver.find_element(By.ID, 'attmRead')
        logging.info("첨부파일 div 찾음: attmRead")

        try:
            size_text = attm_read_div.find_element(By.XPATH, './/span[@class="attm-size"]').text.strip()
            size_match = re.match(r'([\d,\.]+)\s*([KMGT]?B)', size_text, re.IGNORECASE)
            if size_match:
                size_numeric = size_match.group(1).replace(',', '')
                size_unit = size_match.group(2).upper()
                파일용량 = f"{size_numeric} {size_unit}"
            else:
                파일용량 = size_text
            logging.info(f"파일용량 추출: {파일용량}")
        except Exception as e:
            logging.warning(f"파일용량 추출 중 오류 발생: {e}")

        try:
            filename = attm_read_div.find_element(By.XPATH, './/ul[contains(@class, "attm-list")]/li/a/strong').text.strip()
            if '.zip' in filename.lower():
                파일형식 = 'Zip'
            elif '.xlsx' in filename.lower():
                파일형식 = 'Excel'
            logging.info(f"파일형식 추출: {파일형식}")
        except Exception as e:
            logging.warning(f"파일형식 추출 중 오류 발생: {e}")
            파일형식 = ''
    except Exception as e:
        logging.warning(f"attmRead를 찾을 수 없음: {e}")

    if not 파일형식 and not 파일용량:
        try:
            iframe = driver.find_element(By.ID, 'ifa_form')
            driver.switch_to.frame(iframe)
            logging.info("iframe으로 전환하여 파일 정보 추출 시도")
            file_text = find_section_text(driver, ['파밀명 및 용량 (KB)', '파일명 및 용량 (KB)'])
            if file_text:
                logging.info(f"iframe 내에서 파일 정보 추출 시작: {file_text}")
                파일형식, 파일용량 = extract_file_info(file_text)
                logging.info(f"iframe 내에서 파일 정보 추출 완료: {파일형식}, {파일용량}")
            else:
                logging.warning("iframe 내에서 파일 정보 섹션을 찾을 수 없습니다.")
            driver.switch_to.default_content()
        except Exception as e:
            logging.error(f"iframe에서 파일 정보 추출 중 오류 발생: {e}")
            driver.switch_to.default_content()

    return 파일형식, 파일용량

def extract_post_data(driver: webdriver.Chrome, post: webdriver.remote.webelement.WebElement, index: int) -> Optional[Dict]:
    try:
        main_window_handle = driver.current_window_handle
        tds = post.find_elements(By.TAG_NAME, 'td')

        # 등록일
        if len(tds) >= 5:
            등록일_text = tds[4].get_attribute('title').strip() if tds[4].get_attribute('title') else tds[4].text.strip()
        else:
            logging.warning(f"게시글 {index}: 등록일 정보가 부족합니다.")
            등록일_text = ''

        # 작성자
        if len(tds) >= 3:
            작성자_td = tds[2]
            작성자 = 작성자_td.find_element(By.TAG_NAME, 'span').text.strip() if 작성자_td.find_elements(By.TAG_NAME, 'span') else 작성자_td.text.strip()
        else:
            logging.warning(f"게시글 {index}: 작성자 정보가 부족합니다.")
            작성자 = ''

        # 게시글 상세 페이지 이동
        driver.execute_script("arguments[0].scrollIntoView();", post)
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable(post))
        post.click()

        WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(2))
        driver.switch_to.window(driver.window_handles[-1])
        logging.info(f"게시글 {index}: 새 창으로 전환")

        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'HeaderTable')))
        logging.info(f"게시글 {index}: 상세 페이지 로딩 완료")

        제목 = driver.find_element(By.ID, 'DisSubject').text.strip() if driver.find_elements(By.ID, 'DisSubject') else ''
        작성자_full = driver.find_element(By.ID, 'DismyName').text.strip() if driver.find_elements(By.ID, 'DismyName') else ''
        등록일_text_detail = driver.find_element(By.ID, 'DiscDate').text.strip() if driver.find_elements(By.ID, 'DiscDate') else ''

        # 공유대상 추출
        try:
            sharing_target_th = driver.find_element(By.XPATH, '//th[span[contains(text(),"공유대상")]]')
            sharing_target_td = sharing_target_th.find_element(By.XPATH, './following-sibling::td')
            공유대상 = sharing_target_td.get_attribute('textContent').strip()
        except Exception as e:
            logging.info(f"공유대상 추출 실패: {e}")
            공유대상 = ''

        파일형식, 파일용량 = extract_attachment_info(driver)

        # iframe 전환 후 수신자, 개인정보 추출
        법인명, 개인정보_수, 고유식별정보_수, 수신자 = '', 0, 0, ''
        application_form_link = ''
        비고 = ''
        try:
            iframe = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, 'ifa_form'))
            )
            driver.switch_to.frame(iframe)
            logging.info(f"게시글 {index}: iframe으로 전환")

            recipient_text = find_section_text(driver, ['수신자 (부서, 이름)', "Recipient's Department and Name"])
            if recipient_text:
                수신자 = recipient_text.strip()
                logging.info(f"게시글 {index}: 수신자 정보 추출 완료")
            else:
                logging.warning(f"게시글 {index}: 수신자 정보를 찾을 수 없습니다.")

            item_text = find_section_text(driver, ['추출된 항목 및 건수', 'Items and Counts Extracted'])
            if item_text:
                lines = item_text.strip().split('\n')
                keywords = ["주민등록번호", "여권번호", "운전면허의 면허번호", "외국인등록번호", "신분증"]
                found_keywords = False
                for line in lines:
                    line = line.strip()
                    count_match = re.search(r'(\d{1,3}(?:,\d{3})*)\s*건', line)
                    if count_match:
                        count = int(count_match.group(1).replace(',', ''))
                        개인정보_수 += count
                    else:
                        count = 0
                    if any(keyword in line for keyword in keywords):
                        고유식별정보_수 += count
                        found_keywords = True
                if not found_keywords:
                    logging.info(f"게시글 {index}: 고유식별정보 미포함")
                else:
                    logging.info(f"게시글 {index}: 고유식별정보 수 추출 완료: {고유식별정보_수}")
                logging.info(f"게시글 {index}: 개인정보 수 추출 완료: {개인정보_수}")
            else:
                logging.warning(f"게시글 {index}: '추출된 항목 및 건수' 섹션을 찾을 수 없습니다.")

            tr_elements = driver.find_elements(By.XPATH, '//table//tr')
            for tr_elm in tr_elements:
                tds_elm = tr_elm.find_elements(By.TAG_NAME, 'td')
                if len(tds_elm) >= 2:
                    header_text = ''.join([span.text.strip() for span in tds_elm[0].find_elements(By.TAG_NAME, 'span')])
                    if '개인정보 추출 신청서 링크' in header_text or 'URL of the Application Form' in header_text:
                        try:
                            link_element = tds_elm[1].find_element(By.TAG_NAME, 'a')
                            application_form_link = link_element.get_attribute('href')
                            logging.info(f"게시글 {index}: 개인정보 추출 신청서 링크 추출 완료: {application_form_link}")
                        except Exception as e:
                            logging.error(f"게시글 {index}: 개인정보 추출 신청서 링크 추출 중 오류 발생: {e}")
                            cell_text = tds_elm[1].text.strip()
                            if cell_text:
                                application_form_link = cell_text
                                logging.info(f"게시글 {index}: 앵커 태그 없이 일반 텍스트로 링크 대체: {application_form_link}")
                            else:
                                application_form_link = ''
                                logging.info(f"게시글 {index}: 링크 정보 없음")
                        break

            비고_text = find_section_text(driver, ['비고', 'Remark'])
            if 비고_text:
                비고 = 비고_text.strip()

            driver.switch_to.default_content()
        except Exception as e:
            logging.error(f"게시글 {index}: iframe에서 데이터 추출 중 오류 발생: {e}")
            driver.switch_to.default_content()

        # 첨부파일 이력 조회
        진행_구분 = ''
        try:
            attm_log_button = driver.find_element(By.XPATH, '//a[span[text()="첨부파일 이력조회"]]')
            attm_log_button.click()
            logging.info(f"게시글 {index}: 첨부파일 이력조회 버튼 클릭")

            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, '//table[@id="ResultTable"]/tbody/tr'))
                )
                logging.info(f"게시글 {index}: 첨부파일 이력 테이블 로딩 완료")
            except Exception as e:
                logging.error(f"게시글 {index}: 첨부파일 이력 테이블 로딩 중 오류 발생: {e}")
                return None

            rows = driver.find_elements(By.XPATH, '//table[@id="ResultTable"]/tbody/tr')
            다운로드_이력_존재 = False

            for row in rows:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if len(cells) >= 6:
                    구분 = cells[0].text.strip()
                    수행자_element = cells[1]
                    수행자 = 수행자_element.find_element(By.CLASS_NAME, 'pob').text.strip()
                    if 구분 == '다운로드' and 수행자 in 수신자:
                        다운로드_이력_존재 = True
                        logging.info(f"게시글 {index}: 다운로드 이력 발견 - 수행자: {수행자}")
                        break

            if 다운로드_이력_존재:
                진행_구분 = '다운 완료'
            else:
                진행_구분 = ''
        except Exception as e:
            logging.error(f"게시글 {index}: 첨부파일 이력조회 처리 중 오류 발생: {e}")

        등록정보 = 작성자_full
        # 등록정보_부서명 추출
        등록정보_부서명 = extract_department_name(등록정보)

        # 공유대상 처리
        share_segments = [seg.strip() for seg in 공유대상.split(',') if seg.strip()]

        # 수신자와 비교를 통한 best_segment 선택
        def get_word_set(text):
            return set(re.sub(r'[^\w\s]', '', text.replace('님','')).split())

        receiver_set = get_word_set(수신자)
        best_segment = None
        max_overlap = -1

        for seg in share_segments:
            seg_set = get_word_set(seg)
            overlap = len(receiver_set.intersection(seg_set))
            if overlap > max_overlap:
                max_overlap = overlap
                best_segment = seg

        수신자_부서명 = ''
        corp_name = ''
        dept_name = ''
        if best_segment and '/' in best_segment:
            # 예: "유재원 사원/컴투스 별도-세무회계팀"
            parts = best_segment.split('/')
            corp_dept = parts[-1].strip()  # "컴투스 별도-세무회계팀"
            corp_dept_parts = corp_dept.split()
            if len(corp_dept_parts) >= 2:
                corp_name = corp_dept_parts[0].strip()
                dept_name = corp_dept_parts[-1].strip()
            else:
                corp_name = corp_dept_parts[0].strip() if corp_dept_parts else ''
                dept_name = ''

            수신자_부서명 = dept_name
        else:
            # best_segment 없음
            corp_name = ''
            dept_name = ''
            수신자_부서명 = ''

        # 법인명도 여기서 corp_name으로 설정
        법인명 = corp_name

        data = {
            '등록일': 등록일_text or 등록일_text_detail,
            '법인명': 법인명,
            '제목': 제목,
            '작성자': 작성자_full,
            '링크': driver.current_url,
            '파일형식': 파일형식,
            '파일 용량': 파일용량,
            '고유식별정보(수)': 고유식별정보_수,
            '개인정보(수)': 개인정보_수,
            '진행 구분': 진행_구분,
            'application_form_link': application_form_link,
            '등록정보': 등록정보,
            '수신자': 수신자,
            '비고': 비고,
            '등록정보_부서명': 등록정보_부서명,
            '수신자_부서명': 수신자_부서명
        }

        logging.info(f"게시글 {index}: 데이터 추출 완료")
        return data

    except Exception as e:
        logging.error(f"게시글 {index}: 데이터 추출 중 오류 발생: {e}")
        traceback.print_exc()
        return None
    finally:
        try:
            if driver.current_window_handle != main_window_handle:
                driver.close()
                driver.switch_to.window(main_window_handle)
                logging.info(f"게시글 {index}: 새 창 닫기 및 메인 창으로 전환")
            else:
                logging.warning(f"게시글 {index}: 현재 창이 메인 창입니다. 창을 닫지 않습니다.")
            time.sleep(2)
        except Exception as e:
            logging.error(f"창 전환 중 오류 발생: {e}")
            traceback.print_exc()


def save_to_excel(data_list: List[Dict], excel_path: str, driver: webdriver.Chrome) -> None:
    """
    추출된 데이터를 엑셀 파일에 저장하는 함수
    """
    if not data_list:
        logging.info("추출된 데이터가 없습니다.")
        return

    try:
        wb = load_workbook(excel_path)
        if WORKSHEET_NAME not in wb.sheetnames:
            logging.error(f"워크시트 '{WORKSHEET_NAME}'이(가) 존재하지 않습니다.")
            return
        ws = wb[WORKSHEET_NAME]

        # '개인정보 추출 신청' 표 컬럼 매핑 (B~Q)
        request_column_mapping = {
            'NO': 2,
            '결재일': 3,
            '년': 4,
            '월': 5,
            '일': 6,
            '주차': 7,
            '법인명': 8,
            '문서번호': 9,
            '제목': 10,
            '업무 유형': 11,
            '추출 위치': 12,
            '담당 부서': 13,
            '신청자': 14,
            '합의 담당자': 15,
            '링크': 16,
            '진행 구분': 17
        }

        # '개인정보 추출 및 전달' 표 컬럼 매핑 (S~AB)
        delivery_column_mapping = {
            '등록일': 19,
            '법인명': 20,
            '제목': 21,
            '작성자': 22,
            '링크': 23,
            '파일형식': 24,
            '파일 용량': 25,
            '고유식별정보(수)': 26,
            '개인정보(수)': 27,
            '진행 구분': 28
        }

        def find_last_data_row(ws, column_idx, start_row):
            for row_idx in range(ws.max_row, start_row -1, -1):
                cell_value = ws.cell(row=row_idx, column=column_idx).value
                if cell_value is not None and str(cell_value).strip() != '':
                    return row_idx
            return start_row -1

        delivery_start_row = 6
        delivery_link_column = delivery_column_mapping['링크']

        request_start_row = 6
        request_link_column = request_column_mapping['링크']
        request_doc_no_column = request_column_mapping['문서번호']

        last_row_delivery = find_last_data_row(ws, delivery_link_column, delivery_start_row)
        last_row_request = find_last_data_row(ws, request_link_column, request_start_row)

        max_row = max(last_row_delivery, last_row_request)

        # NO 부여를 위한 next_no 계산
        last_no = None
        for r in range(ws.max_row, request_start_row-1, -1):
            val = ws.cell(row=r, column=request_column_mapping['NO']).value
            if val and isinstance(val, int):
                last_no = val
                break
        if last_no is None:
            next_no = 1
        else:
            next_no = last_no + 1

        for data in data_list:
            application_form_link = data.get('application_form_link', '').strip()
            delivery_row_idx = None

            found_row_in_request = None
            for row_idx in range(request_start_row, last_row_request + 1):
                cell_value = ws.cell(row=row_idx, column=request_link_column).value
                if cell_value and str(cell_value).strip() == application_form_link:
                    found_row_in_request = row_idx
                    break

            if not found_row_in_request and 'gw.com2us.com' in application_form_link:
                try:
                    driver.get(application_form_link)
                    time.sleep(1)
                    try:
                        expected_link_element = driver.find_element(By.LINK_TEXT, "예상되는 완료함 문서로 이동")
                        if expected_link_element:
                            expected_link_element.click()
                            time.sleep(1)
                    except:
                        pass

                    doc_no = ''
                    try:
                        doc_no_elements = driver.find_elements(By.XPATH, '//th[contains(text(),"문서번호")]/following-sibling::td[1]')
                        doc_no = doc_no_elements[0].text.strip() if doc_no_elements else ''
                        logging.info(f"추출한 문서번호: {doc_no}")
                    except Exception as e:
                        logging.error(f"문서번호 추출 중 오류 발생: {e}")

                    if doc_no:
                        for row_idx in range(request_start_row, last_row_request + 1):
                            cell_value = ws.cell(row=row_idx, column=request_doc_no_column).value
                            if cell_value and str(cell_value).strip() == doc_no:
                                found_row_in_request = row_idx
                                logging.info(f"문서번호 일치: 행 {row_idx}")
                                break

                except Exception as e:
                    logging.error(f"application_form_link 확인 중 오류 발생: {e}")

            if found_row_in_request:
                delivery_row_idx = found_row_in_request
                for col_name, col_idx in delivery_column_mapping.items():
                    if col_name in data:
                        ws.cell(row=delivery_row_idx, column=col_idx, value=data[col_name])
                logging.info(f"데이터가 '개인정보 추출 및 전달' 표의 행 {delivery_row_idx}에 저장되었습니다.")
            else:
                delivery_row_idx = max_row + 1
                max_row += 1
                for col_name, col_idx in delivery_column_mapping.items():
                    if col_name in data:
                        ws.cell(row=delivery_row_idx, column=col_idx, value=data[col_name])
                logging.info(f"데이터가 '개인정보 추출 및 전달' 표의 새로운 행 {delivery_row_idx}에 저장되었습니다.")
                last_row_delivery = delivery_row_idx

            keywords = ["구글폼", "구글 폼", "네이버폼", "네이버 폼"]
            form_condition = (any(kw in application_form_link for kw in keywords) or
                              '@' in application_form_link or
                              'forms' in application_form_link.lower() or
                              'form' in application_form_link.lower())

            if form_condition:
                # 같은 행 번호에 추가
                request_row_idx = delivery_row_idx

                # NO 부여
                if ws.cell(row=request_row_idx, column=request_column_mapping['NO']).value is None:
                    ws.cell(row=request_row_idx, column=request_column_mapping['NO'], value=next_no)
                    next_no += 1

                등록일 = data.get('등록일', '')
                법인명 = data.get('법인명', '')
                작성자 = data.get('작성자', '')
                진행_구분 = '추출완료'
                등록정보_부서명 = data.get('등록정보_부서명', '')
                수신자_부서명 = data.get('수신자_부서명', '')

                # 여기서 결재일은 data['등록일']을 사용
                결재일_value = 등록일  # data['등록일']가 YYYY-MM-DD 형태라고 가정
                년_val, 월_val, 일_val = '', '', ''
                if 결재일_value and '-' in 결재일_value:
                    parts = 결재일_value.split('-')
                    if len(parts) == 3:
                        년_val = parts[0]
                        월_val = str(int(parts[1]))
                        일_val = str(int(parts[2]))

                제목 = f"{등록정보_부서명} 직접 수집 및 {수신자_부서명} 전달 건"

                비고 = data.get('비고', '')
                추출위치 = ''
                if '@' in application_form_link:
                    추출위치 = '메일'
                elif any(kw in application_form_link for kw in keywords):
                    추출위치 = '외부폼'
                elif 'forms' in application_form_link.lower() or 'form' in application_form_link.lower():
                    추출위치 = '외부폼'
                elif any(kw in 비고 for kw in keywords):
                    추출위치 = '외부폼'

                logging.info(f"등록일: {등록일}, 법인명: {법인명}, 작성자: {작성자}")
                logging.info(f"등록정보_부서명: {등록정보_부서명}, 수신자_부서명: {수신자_부서명}")
                logging.info(f"생성된 제목: {제목}")
                logging.info(f"application_form_link: {application_form_link}, 비고: {비고}, 추출위치 결정: {추출위치}")

                # 데이터 입력 (NO, 결재일, 년, 월, 일)
                ws.cell(row=request_row_idx, column=request_column_mapping['결재일'], value=결재일_value)
                ws.cell(row=request_row_idx, column=request_column_mapping['년'], value=년_val)
                ws.cell(row=request_row_idx, column=request_column_mapping['월'], value=월_val)
                ws.cell(row=request_row_idx, column=request_column_mapping['일'], value=일_val)
                ws.cell(row=request_row_idx, column=request_column_mapping['법인명'], value=법인명)
                ws.cell(row=request_row_idx, column=request_column_mapping['문서번호'], value='-')
                ws.cell(row=request_row_idx, column=request_column_mapping['제목'], value=제목)
                ws.cell(row=request_row_idx, column=request_column_mapping['업무 유형'], value='')
                ws.cell(row=request_row_idx, column=request_column_mapping['추출 위치'], value=추출위치)
                ws.cell(row=request_row_idx, column=request_column_mapping['담당 부서'], value=등록정보_부서명)
                ws.cell(row=request_row_idx, column=request_column_mapping['신청자'], value=작성자)
                ws.cell(row=request_row_idx, column=request_column_mapping['합의 담당자'], value='-')
                ws.cell(row=request_row_idx, column=request_column_mapping['링크'], value='-')
                ws.cell(row=request_row_idx, column=request_column_mapping['진행 구분'], value=진행_구분)
                logging.info(f"데이터가 '개인정보 추출 신청' 표의 행 {request_row_idx}에 저장되었습니다.")
            else:
                logging.info("조건에 맞지 않아 '개인정보 추출 신청' 표에 데이터를 추가하지 않습니다.")

        try:
            max_row = max(last_row_delivery, last_row_request)
            for row_idx in range(6, max_row + 1):
                cell_value = ws.cell(row=row_idx, column=delivery_column_mapping['제목']).value
                if cell_value and '추출완료' in str(cell_value):
                    ws.cell(row=row_idx, column=request_column_mapping['진행 구분'], value='추출 완료')
                    logging.info(f"행 {row_idx}: '추출 완료'를 Column Q에 저장했습니다.")
        except Exception as e:
            logging.error(f"마지막 단계에서 오류 발생: {e}")

        wb.save(excel_path)
        logging.info(f"데이터가 성공적으로 '{excel_path}' 파일에 저장되었습니다.")

    except Exception as e:
        logging.error("엑셀 파일 처리 중 오류가 발생했습니다.")
        logging.error(e)
        traceback.print_exc()


def main(username: str, password: str, max_posts: Optional[int] = None) -> Optional[str]:
    if max_posts is None:
        max_posts = MAX_POSTS

    driver = initialize_webdriver()

    try:
        login_attempts = 0
        max_login_attempts = 3

        while login_attempts < max_login_attempts:
            if login(driver, username, password):
                break
            else:
                login_attempts += 1
                logging.error("로그인에 실패하였습니다. 다시 시도하세요.")
                username = input("사용자명을 입력하세요: ")
                password = getpass.getpass("비밀번호를 입력하세요: ")
                driver.get('https://gw.com2us.com/')
        else:
            logging.error("로그인 시도 횟수를 초과하였습니다.")
            driver.quit()
            return None

        if not navigate_to_target_page(driver):
            driver.quit()
            return None

        data_list = []
        total_crawled = 0
        page_number = 1

        while total_crawled < max_posts:
            logging.info(f"{page_number} 페이지 크롤링 시작")
            posts = fetch_posts(driver)
            if not posts:
                logging.info(f"{page_number} 페이지에 처리할 게시글이 없습니다.")
                break

            if page_number == 1 and len(posts) > 0:
                posts = posts[1:]

            num_posts_to_crawl = min(len(posts), max_posts - total_crawled)

            for i in range(num_posts_to_crawl):
                if total_crawled >= max_posts:
                    break
                if i >= len(posts):
                    logging.warning(f"게시글 수가 예상보다 적습니다. 현재 인덱스: {i}, 게시글 수: {len(posts)}")
                    break

                data = extract_post_data(driver, posts[i], total_crawled + 1)
                if data:
                    logging.info(f"data 딕셔너리 내용: {data}")
                    data_list.append(data)
                    total_crawled += 1

            if total_crawled >= max_posts:
                break

            page_number += 1
            if not go_to_page(driver, page_number):
                logging.info("더 이상 페이지가 없습니다.")
                break

        save_to_excel(data_list, EXCEL_FILE, driver)

        logging.info("개인정보 추출 및 전달이 완료되었습니다.")

        run_extraction_in_progress(driver)

        return EXCEL_FILE

    except Exception as e:
        logging.error("스크립트 실행 중 예상치 못한 오류가 발생했습니다.")
        logging.error(e)
        traceback.print_exc()
        return None
    finally:
        driver.quit()
        logging.info("브라우저가 종료되었습니다.")


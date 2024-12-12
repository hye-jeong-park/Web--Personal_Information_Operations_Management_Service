import time
import traceback
import logging
import datetime
from typing import Optional, List, Dict

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# 크롤링할 최대 게시글 수 설정
MAX_POSTS = 50
EXCEL_FILE = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'  # 엑셀 파일 경로
WORKSHEET_NAME = '개인정보 추출 및 이용 관리'

def navigate_to_search_page(driver) -> bool:
    """
    검색 페이지로 이동하는 함수
    """
    try:
        # 검색 페이지 URL로 이동
        driver.get('https://gw.com2us.com/emate_appro/appro_link.nsf/view?readform&viewname=view04&vctype=a')
        # 검색창이 로드될 때까지 대기
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'searchtext')))
        return True
    except Exception as e:
        logging.error(f"검색 페이지 이동 중 오류 발생: {e}")
        traceback.print_exc()
        return False

def search_documents(driver) -> bool:
    """
    '개인정보 추출 신청서' 검색을 수행하는 함수
    """
    try:
        # 검색어 입력
        search_input = driver.find_element(By.ID, 'searchtext')
        search_input.clear()
        search_input.send_keys('개인정보 추출 신청서')
        
        # 검색 버튼 클릭
        search_button = driver.find_element(By.XPATH, '//img[@class="inbtn" and contains(@src, "btn_search_board.gif")]')
        search_button.click()
        time.sleep(5)
        return True
    except Exception as e:
        logging.error(f"문서 검색 중 오류 발생: {e}")
        traceback.print_exc()
        return False

def extract_post_data(driver, post, index: int) -> Optional[Dict]:
    """
    개별 게시글의 데이터를 추출하는 함수
    """
    try:
        # 기본 정보 추출
        tds = post.find_elements(By.TAG_NAME, 'td')
        
        # 결재일 추출
        결재일_text = tds[5].text.strip()
        
        # 날짜 문자열에서 날짜 부분만 추출
        try:
            # 결재일_text를 공백으로 분리하여 첫 번째 요소를 가져옴
            date_part = 결재일_text.split()[0]
            # 날짜 문자열을 datetime 객체로 변환
            date_obj = datetime.datetime.strptime(date_part, '%Y-%m-%d')
        except ValueError:
            logging.error(f"결재일 형식이 올바르지 않습니다: {결재일_text}")
            return None
        
        년 = str(date_obj.year)
        월 = str(date_obj.month)
        일 = str(date_obj.day)
        
        # 신청자 추출
        신청자 = tds[4].find_element(By.TAG_NAME, 'span').text.strip()

        # 게시글 상세 페이지 열기
        driver.execute_script("arguments[0].scrollIntoView();", post)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable(post))
        post.click()

        # 새 창으로 전환
        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
        driver.switch_to.window(driver.window_handles[-1])

        # 상세 페이지 데이터 추출
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'AppLineArea')))
        
        # 문서 종류 확인
        h2_element = driver.find_element(By.CSS_SELECTOR, '#AppLineArea h2')
        if h2_element.text.strip() != '개인정보 추출 신청서':
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            return None

        # 상세 정보 추출
        법인명_elements = driver.find_elements(By.ID, 'titleLabel')
        법인명 = 법인명_elements[0].text.strip() if 법인명_elements else ''

        문서번호_elements = driver.find_elements(By.XPATH, '//th[contains(text(),"문서번호")]/following-sibling::td[1]')
        문서번호 = 문서번호_elements[0].text.strip() if 문서번호_elements else ''

        제목_elements = driver.find_elements(By.CSS_SELECTOR, 'td.approval_text')
        if 제목_elements:
            제목_text = 제목_elements[0].text.strip()
            # 제목이 법인명 + 공백으로 시작하면, 그 부분을 제거
            if 제목_text.startswith(법인명 + ' '):
                제목 = 제목_text[len(법인명) + 1:].strip()
            else:
                제목 = 제목_text
        else:
            제목 = ''

        합의담당자_elements = driver.find_elements(By.XPATH, '//th[text()="합의선"]/following::tr[@class="name"][1]/td[@class="td_point"]')
        합의담당자 = 합의담당자_elements[0].text.strip() if 합의담당자_elements else ''

        # 업무유형 추출
        업무유형 = ''
        추출위치 = 'DB'  # 추출위치는 'DB'로 설정
        담당부서 = ''

        try:
            # iframe으로 전환
            iframe = driver.find_element(By.ID, 'ifa_form')
            driver.switch_to.frame(iframe)
            logging.info("iframe으로 전환하여 업무유형 및 담당부서 추출 시작")

            # '구분'이라는 텍스트를 포함하는 <th>를 찾음
            gu_bun_th = driver.find_element(By.XPATH, '//th[span[contains(text(), "구분")]]')
            # '구분' <th>의 부모 <tr>을 찾음
            gu_bun_tr = gu_bun_th.find_element(By.XPATH, './ancestor::tr')
            # '구분' <tr>과 그 다음 모든 형제 <tr>들에서 체크박스와 라벨을 추출
            checkbox_trs = [gu_bun_tr] + gu_bun_tr.find_elements(By.XPATH, './following-sibling::tr')

            for tr in checkbox_trs:
                # 해당 <tr> 안의 모든 체크박스들을 찾음
                checkboxes = tr.find_elements(By.XPATH, './/input[@type="checkbox"]')
                for checkbox in checkboxes:
                    # 체크박스가 선택되었는지 확인
                    is_checked = checkbox.is_selected()
                    # 체크박스 바로 다음의 형제 노드에서 라벨(span)을 찾음
                    label = checkbox.find_element(By.XPATH, './following-sibling::span[1]')
                    label_text = label.text.strip()
                    logging.info(f"라벨: {label_text}, 선택됨: {is_checked}")
                    if is_checked:
                        if label_text == '프로모션 관리(사전등록, 각 종 이벤트)':
                            업무유형 = '사전예약/이벤트'
                            break
                        elif label_text == '미접속 사용자 대상 이벤트':
                            업무유형 = '홍보/광고'
                            break
                        elif label_text == '통신비밀 보호업무 요청':
                            업무유형 = '통비'
                            break
                        elif label_text == '기타':
                            업무유형 = '기타'
                            break
                if 업무유형:
                    break

            logging.info(f"업무유형 추출 완료: {업무유형}")

            # 담당부서 추출
            department_td = driver.find_element(By.XPATH, '//th[span[text()="부서"]]/following-sibling::td')
            담당부서_full = department_td.text.strip()
            담당부서 = 담당부서_full.split()[-1] if 담당부서_full else ''
            logging.info(f"담당부서 추출 완료: {담당부서}")

            driver.switch_to.default_content()
        except Exception as e1:
            if 'no such element: Unable to locate element' in str(e1):
                logging.warning(f"첫 번째 방법 실패: {e1}. 두 번째 방법으로 재시도합니다.")
                try:
                    # '구분'이라는 텍스트를 포함하는 <th>를 찾음
                    gu_bun_th = driver.find_element(By.XPATH, '//th[contains(text(), "구분")]')
                    # '구분' <th>의 부모 <tr>을 찾음
                    gu_bun_tr = gu_bun_th.find_element(By.XPATH, './ancestor::tr')

                    # '구분' 섹션에 속하는 모든 <tr> 요소를 수집
                    checkbox_trs = []
                    current_tr = gu_bun_tr
                    while True:
                        checkbox_trs.append(current_tr)
                        try:
                            # 다음 형제 <tr>을 찾음
                            next_tr = current_tr.find_element(By.XPATH, './following-sibling::tr[1]')
                            # 다음 <tr>에 <th> 요소가 있는지 확인
                            th_in_next_tr = next_tr.find_elements(By.XPATH, './/th')
                            if th_in_next_tr:
                                # <th> 요소가 있으면 다른 섹션이므로 종료
                                break
                            else:
                                # <th> 요소가 없으면 계속 탐색
                                current_tr = next_tr
                        except:
                            # 더 이상 <tr>이 없으면 종료
                            break
                    
                    # 수집된 <tr>들에서 체크박스와 라벨 추출
                    업무유형 = ''
                    for tr in checkbox_trs:
                        checkboxes = tr.find_elements(By.XPATH, './/input[@type="checkbox"]')
                        for checkbox in checkboxes:
                            is_checked = checkbox.is_selected()

                            # 체크박스의 다음 텍스트 노드의 내용을 가져옴
                            label_text = driver.execute_script('''
                                var node = arguments[0].nextSibling;
                                while(node && (node.nodeType !== Node.TEXT_NODE || !node.textContent.trim())) {
                                    node = node.nextSibling;
                                }
                                return node ? node.textContent.trim() : '';
                            ''', checkbox)

                            logging.info(f"라벨: {label_text}, 선택됨: {is_checked}")
                            if is_checked:
                                if label_text == '프로모션 관리(사전등록, 각 종 이벤트)':
                                    업무유형 = '사전예약/이벤트'
                                    break
                                elif label_text == '미접속 사용자 대상 이벤트':
                                    업무유형 = '홍보/광고'
                                    break
                                elif label_text == '통신비밀 보호업무 요청':
                                    업무유형 = '통비'
                                    break
                                elif label_text == '기타':
                                    업무유형 = '기타'
                                    break
                        if 업무유형:
                            break

                    logging.info(f"두 번째 방법으로 업무유형 추출 완료: {업무유형}")

                    # 담당부서 추출
                    department_td = driver.find_element(By.XPATH, '//th[text()="부서"]/following-sibling::td')
                    담당부서 = department_td.text.strip()
                    logging.info(f"두 번째 방법으로 담당부서 추출 완료: {담당부서}")

                    driver.switch_to.default_content()
                except Exception as e2:
                    logging.error(f"두 번째 방법도 실패하였습니다: {e2}")
                    업무유형 = ''
                    담당부서 = ''
                    driver.switch_to.default_content()
            else:
                # 다른 예외의 경우 기존 처리를 유지
                logging.error(f"업무유형 및 담당부서 추출 중 예기치 않은 오류 발생: {e1}")
                업무유형 = ''
                담당부서 = ''
                driver.switch_to.default_content()
                         
        # 추출 데이터 구성
        data = {
            '결재일': 결재일_text,
            '년': 년,
            '월': 월,
            '일': 일,
            '주차': '',
            '법인명': 법인명,
            '문서번호': 문서번호,
            '제목': 제목,
            '업무 유형': 업무유형,
            '추출 위치': 추출위치,
            '담당 부서': 담당부서,
            '신청자': 신청자,
            '합의 담당자': 합의담당자,
            '링크': driver.current_url,
            '진행 구분': ''
        }

        logging.info(f"게시글 {index}: 데이터 추출 완료")
        return data

    except Exception as e:
        logging.error(f"게시글 {index}: 데이터 추출 중 오류 발생: {e}")
        traceback.print_exc()
        return None
    finally:
        try:
            window_handles = driver.window_handles
            if len(window_handles) > 1:
                # 현재 창(팝업 창)을 닫고, 메인 창으로 전환
                driver.close()
                driver.switch_to.window(window_handles[0])
            else:
                # 창이 하나밖에 없을 경우, 전환만 수행
                driver.switch_to.window(window_handles[0])
            time.sleep(2)
        except Exception as e:
            logging.error(f"창 전환 중 오류 발생: {e}")
            traceback.print_exc()

            
def save_to_excel(data_list: List[Dict]) -> None:
    if not data_list:
        logging.info("추출된 데이터가 없습니다.")
        return

    try:
        # 엑셀 파일 열기
        wb = load_workbook(EXCEL_FILE)
        if WORKSHEET_NAME not in wb.sheetnames:
            logging.error(f"워크시트 '{WORKSHEET_NAME}'이(가) 존재하지 않습니다.")
            return
        ws = wb[WORKSHEET_NAME]

        # 데이터가 시작되는 행 번호 (예: 6행)
        data_start_row = 6

        # 마지막 행 찾기
        last_row = ws.max_row

        # 시작 행 설정
        if last_row < data_start_row:
            start_row = data_start_row
        else:
            # 마지막 행에 데이터가 있는지 확인
            while last_row >= data_start_row:
                cell_values = [ws.cell(row=last_row, column=col_idx).value for col_idx in range(2, 29)]  # B~AC열 검사
                if any(cell_values):
                    break
                last_row -= 1
            start_row = last_row + 1

        # 엑셀 열 매핑 정의 (B~Q열에 데이터 저장)
        column_mapping = {
            '결재일': 3,    # C열
            '년': 4,        # D열
            '월': 5,        # E열
            '일': 6,        # F열
            '주차': 7,      # G열
            '법인명': 8,     # H열
            '문서번호': 9,   # I열
            '제목': 10,      # J열
            '업무 유형': 11,  # K열
            '추출 위치': 12,  # L열
            '담당 부서': 13,  # M열
            '신청자': 14,     # N열
            '합의 담당자': 15, # O열
            '링크': 16,       # P열
            '진행 구분': 17   # Q열
        }

        # 데이터 쓰기
        for data in data_list:
            ws.cell(row=start_row, column=2, value=start_row - data_start_row + 1)  # 번호 저장 (B열)
            for col_name, col_idx in column_mapping.items():
                ws.cell(row=start_row, column=col_idx, value=data.get(col_name, ''))
            start_row += 1

        # 파일 저장
        wb.save(EXCEL_FILE)
        logging.info(f"데이터가 성공적으로 '{EXCEL_FILE}' 파일에 저장되었습니다.")

    except Exception as e:
        logging.error(f"엑셀 저장 중 오류 발생: {e}")
        traceback.print_exc()
        

def run_extraction_in_progress(driver, max_posts: Optional[int] = None):
    """
    메인 함수
    """
    try:
        if not navigate_to_search_page(driver):
            return None

        if not search_documents(driver):
            return None

        # 게시글 데이터 추출
        data_list = []
        posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')

        # 최대 게시글 수 설정
        if max_posts is not None:
            num_posts_to_crawl = min(len(posts), max_posts)
        else:
            num_posts_to_crawl = len(posts)

        for i in range(num_posts_to_crawl):
            if i >= len(posts):
                logging.warning(f"게시글 수가 예상보다 적습니다. 현재 인덱스: {i}, 게시글 수: {len(posts)}")
                break

            data = extract_post_data(driver, posts[i], i + 1)
            if data:
                data_list.append(data)

        # 추출된 데이터 저장
        save_to_excel(data_list)

        # 완료 메시지와 파일 경로 반환
        logging.info("개인정보 신청 이력 저장이 완료되었습니다.")
        return EXCEL_FILE

    except Exception as e:
        logging.error(f"메인 프로세스 중 오류 발생: {e}")
        traceback.print_exc()
        return None

# 스크립트를 직접 실행할 경우를 대비하여 추가
if __name__ == '__main__':
    logging.error("이 스크립트는 단독으로 실행될 수 없습니다.")
    logging.error("delivery_script.py를 통해 실행해 주세요.")




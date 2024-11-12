##저장할 파일 선택하고(구현 완료), 링크 통해서 파일 다운로드하는 방법(추가 구현 필요)##
import sys
import time
import traceback
import logging
from typing import Optional, List, Dict

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def initialize_webdriver() -> webdriver.Chrome:
    """
    웹드라이버 초기화
    """
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    return driver


def login(driver: webdriver.Chrome, username: str, password: str) -> bool:
    """
    로그인 처리
    """
    try:
        driver.get('https://gw.com2us.com/')
        username_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'Username'))
        )
        password_input = driver.find_element(By.ID, 'Password')
        username_input.send_keys(username)
        password_input.send_keys(password)
        driver.find_element(By.CLASS_NAME, 'btnLogin').click()

        WebDriverWait(driver, 20).until(EC.url_changes('https://gw.com2us.com/'))
        if 'login' in driver.current_url.lower():
            logging.error("로그인에 실패하였습니다.")
            return False
        return True
    except Exception as e:
        logging.error(f"로그인 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def navigate_to_search_page(driver: webdriver.Chrome) -> bool:
    """
    검색 페이지로 이동하는 함수
    """
    try:
        driver.get('https://gw.com2us.com/emate_appro/appro_complete_2024_link.nsf/wfmViaView?readform&viewname=view055&vctype=a')
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'searchtext')))
        return True
    except Exception as e:
        logging.error(f"검색 페이지 이동 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def search_documents(driver: webdriver.Chrome) -> bool:
    """
    '개인정보 추출 신청서' 검색을 수행하는 함수
    """
    try:
        search_input = driver.find_element(By.ID, 'searchtext')
        search_input.clear()
        search_input.send_keys('개인정보 추출 신청서')

        search_button = driver.find_element(By.XPATH, '//img[@class="inbtn" and contains(@src, "btn_search_board.gif")]')
        search_button.click()
        time.sleep(5)
        return True
    except Exception as e:
        logging.error(f"문서 검색 중 오류 발생: {e}")
        traceback.print_exc()
        return False


def extract_post_data(driver: webdriver.Chrome, post: webdriver.remote.webelement.WebElement, index: int) -> Optional[Dict]:
    """
    개별 게시글의 데이터를 추출하는 함수
    """
    try:
        tds = post.find_elements(By.TAG_NAME, 'td')
        결재일_text = tds[5].text.strip()
        년, 월, 일 = 결재일_text.split('-')
        월 = str(int(월))
        일 = str(int(일))

        신청자 = tds[4].find_element(By.TAG_NAME, 'span').text.strip()

        driver.execute_script("arguments[0].scrollIntoView();", post)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable(post))
        post.click()

        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
        driver.switch_to.window(driver.window_handles[-1])

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'AppLineArea')))

        h2_element = driver.find_element(By.CSS_SELECTOR, '#AppLineArea h2')
        if h2_element.text.strip() != '개인정보 추출 신청서':
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            return None

        법인명_elements = driver.find_elements(By.ID, 'titleLabel')
        법인명 = 법인명_elements[0].text.strip() if 법인명_elements else ''

        문서번호_elements = driver.find_elements(By.XPATH, '//th[contains(text(),"문서번호")]/following-sibling::td[1]')
        문서번호 = 문서번호_elements[0].text.strip() if 문서번호_elements else ''

        제목_elements = driver.find_elements(By.CSS_SELECTOR, 'td.approval_text')
        if 제목_elements:
            제목_text = 제목_elements[0].text.strip()
            제목 = 제목_text[len(법인명) + 1:].strip() if 제목_text.startswith(법인명 + ' ') else 제목_text
        else:
            제목 = ''

        합의담당자_elements = driver.find_elements(By.XPATH, '//th[text()="합의선"]/following::tr[@class="name"][1]/td[@class="td_point"]')
        합의담당자 = 합의담당자_elements[0].text.strip() if 합의담당자_elements else ''

        data = {
            '결재일': 결재일_text,
            '년': 년,
            '월': 월,
            '일': 일,
            '주차': '',
            '법인명': 법인명,
            '문서번호': 문서번호,
            '제목': 제목,
            '업무 유형': '',
            '추출 위치': '',
            '담당 부서': '',
            '신청자': 신청자,
            '합의 담당자': 합의담당자,
            '링크': driver.current_url,
            '진행 구분': ''
        }

        logging.info(f"게시글 {index}: 데이터 추출 완료")
        return data

    except Exception as e:
        logging.error(f"게시글 {index}: 데이터 추출 중 오류 발생: {e}")
        traceback.print_exc()
        return None
    finally:
        try:
            window_handles = driver.window_handles
            if len(window_handles) > 1:
                driver.close()
                driver.switch_to.window(window_handles[0])
            else:
                driver.switch_to.window(window_handles[0])
            time.sleep(2)
        except Exception as e:
            logging.error(f"창 전환 중 오류 발생: {e}")
            traceback.print_exc()


def save_to_excel(data_list: List[Dict], excel_path: str) -> None:
    """
    추출된 데이터를 엑셀 파일에 저장하는 함수
    """
    if not data_list:
        logging.info("추출된 데이터가 없습니다.")
        return

    try:
        wb = load_workbook(excel_path)
        ws = wb['개인정보 추출 및 이용 관리']

        last_row = ws.max_row
        while last_row >= 6:
            if ws.cell(row=last_row, column=2).value is not None:
                break
            last_row -= 1

        start_row = max(6, last_row + 1)
        next_no = ws.cell(row=last_row, column=2).value + 1 if last_row >= 6 else 1

        column_mapping = {
            '결재일': 3, '년': 4, '월': 5, '일': 6, '주차': 7,
            '법인명': 8, '문서번호': 9, '제목': 10, '업무 유형': 11,
            '추출 위치': 12, '담당 부서': 13, '신청자': 14,
            '합의 담당자': 15, '링크': 16, '진행 구분': 17
        }

        for data in data_list:
            ws.cell(row=start_row, column=2, value=next_no)
            for col_name, col_idx in column_mapping.items():
                ws.cell(row=start_row, column=col_idx, value=data[col_name])
            start_row += 1
            next_no += 1

        wb.save(excel_path)
        logging.info(f"데이터가 성공적으로 '{excel_path}' 파일에 저장되었습니다.")

    except Exception as e:
        logging.error(f"엑셀 저장 중 오류 발생: {e}")
        traceback.print_exc()


def main(username: str, password: str, max_posts: Optional[int], excel_path: str) -> Optional[str]:
    driver = initialize_webdriver()
    try:
        if not login(driver, username, password):
            return None
        if not navigate_to_search_page(driver):
            return None
        if not search_documents(driver):
            return None

        data_list = []
        posts = driver.find_elements(By.XPATH, '//tr[contains(@class, "dhx_skyblue")]')
        num_posts_to_crawl = min(len(posts), max_posts) if max_posts else len(posts)

        for i in range(num_posts_to_crawl):
            posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
            if i >= len(posts):
                logging.warning(f"게시글 수가 예상보다 적습니다. 현재 인덱스: {i}, 게시글 수: {len(posts)}")
                break

            data = extract_post_data(driver, posts[i], i + 1)
            if data:
                data_list.append(data)

        save_to_excel(data_list, excel_path)
        logging.info("개인정보 신청 이력 저장이 완료되었습니다.")
        return excel_path

    except Exception as e:
        logging.error(f"메인 프로세스 중 오류 발생: {e}")
        traceback.print_exc()
        return None
    finally:
        driver.quit()
        logging.info("브라우저가 종료되었습니다.")
